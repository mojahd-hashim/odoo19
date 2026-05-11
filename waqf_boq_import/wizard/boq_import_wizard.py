import base64
import json
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError


class BOQImportWizard(models.TransientModel):
    _name = 'mosque.boq.import.wizard'
    _description = 'BOQ Import Wizard — from Excel or JSON'

    import_mode = fields.Selection([
        ('json',  'JSON File (pre-processed)'),
        ('excel', 'Excel File (raw BOQ)'),
        ('all',   'Load All — from embedded data'),
    ], string='Import Mode', default='all', required=True)

    file_data = fields.Binary(string='File to Import')
    file_name = fields.Char(string='File Name')

    overwrite = fields.Boolean(
        string='Overwrite existing BOQ lines',
        default=False,
        help='If checked, existing BOQ lines for the mosque will be deleted first.',
    )
    mosque_ids = fields.Many2many(
        'mosque.mosque',
        string='Target Mosques',
        help='Leave empty to import for all mosques found in the file.',
    )

    result_summary = fields.Text(string='Import Result', readonly=True)
    state = fields.Selection([
        ('draft',  'Ready'),
        ('done',   'Done'),
        ('error',  'Error'),
    ], default='draft')

    # ── Category resolution ───────────────────────────────────────
    def _get_category(self, code):
        Cat = self.env['mosque.boq.category']
        cat = Cat.search([('code', '=', code)], limit=1)
        if not cat:
            # Fallback: create on the fly
            type_map = {'ARCH': 'architectural', 'MECH': 'mechanical', 'ELEC': 'electrical'}
            name_map = {
                'ARCH': 'Architectural & Structural Works',
                'MECH': 'Mechanical Works',
                'ELEC': 'Electrical Works',
            }
            cat = Cat.create({
                'code': code,
                'name': name_map.get(code, code),
                'type': type_map.get(code, 'architectural'),
            })
        return cat

    # ── Mosque resolution ─────────────────────────────────────────
    def _get_mosque(self, code):
        return self.env['mosque.mosque'].search(
            [('code', '=', code)], limit=1)

    # ── Main import action ────────────────────────────────────────
    def action_import(self):
        self.ensure_one()

        if self.import_mode == 'all':
            return self._import_embedded()
        elif self.import_mode == 'json':
            return self._import_json()
        elif self.import_mode == 'excel':
            return self._import_excel()

    # ── Mode 1: embedded data (the actual BOQ from the contract) ──
    def _import_embedded(self):
        """
        Import from the embedded BOQ data extracted from the Excel file.
        This is the fastest path — all 1831 lines are baked in.
        """
        from odoo.addons.waqf_boq_import.data.boq_embedded import BOQ_DATA
        return self._process_rows(BOQ_DATA)

    # ── Mode 2: JSON file ─────────────────────────────────────────
    def _import_json(self):
        if not self.file_data:
            raise UserError(_('Please upload a JSON file.'))
        raw = base64.b64decode(self.file_data)
        try:
            data = json.loads(raw.decode('utf-8'))
        except Exception as e:
            raise UserError(_('Invalid JSON file: %s') % str(e))
        return self._process_rows(data)

    # ── Mode 3: Excel file ────────────────────────────────────────
    def _import_excel(self):
        if not self.file_data:
            raise UserError(_('Please upload the BOQ Excel file.'))
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('openpyxl is required. Run: pip install openpyxl'))

        raw = base64.b64decode(self.file_data)
        wb  = load_workbook(io.BytesIO(raw), read_only=True)

        SHEET_TO_MOSQUE = {
            '1':  'RUH-16', '2':  'RUH-17', '3':  'RUH-09',
            '4':  'RUH-18', '5':  'RUH-20', '6':  'RUH-06',
            '7':  'RUH-07', '8':  'RUH-12', '9':  'RUH-08',
            '10': 'RUH-13', '11': 'RUH-11', '12': 'JED-01',
            '13': 'JED-02', '14': 'TIF-01', '15': 'TIF-02',
            '16': 'RUH-01', '17': 'RUH-10', '18': 'RUH-03',
            '19': 'RUH-05', '20': 'RUH-02', '21': 'RUH-04',
            '22': 'RFH-01', '23': 'RFH-02', '24': 'RUH-15',
            '25': 'RUH-14', '26': 'RUH-19', '27': 'AFJ-01',
            '28': 'YRA-01', '29': 'GIZ-02',
        }

        def detect_category(t):
            t = str(t or '')
            if 'كهرب' in t:                                    return 'ELEC'
            if 'ميكانيك' in t or 'صحي' in t or 'تكييف' in t: return 'MECH'
            return 'ARCH'

        def map_uom(t):
            t = str(t or '').strip()
            if 'متر طولي' in t or 'م ط' in t: return 'm'
            if 'متر مسطح' in t or 'م2' in t:  return 'm2'
            if 'متر مكعب' in t or 'م3' in t:  return 'm3'
            if 'مقطوع' in t or 'مجموع' in t:  return 'ls'
            return 'unit'

        rows = []
        for sheet_name in wb.sheetnames:
            if sheet_name.strip() == 'الاجمالى':
                continue
            mosque_code = SHEET_TO_MOSQUE.get(sheet_name)
            if not mosque_code:
                continue
            ws = wb[sheet_name]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 9:
                    continue
                if row[0] is None:
                    continue
                try:
                    item_num = row[0]
                    if not isinstance(item_num, (int, float)):
                        continue
                    cat_text  = row[1]
                    item_name = str(row[2] or '').strip()
                    uom_text  = row[4]
                    qty       = row[5]
                    price     = row[6]
                    if not item_name or qty is None or price is None:
                        continue
                    if 'مجموع' in item_name or 'إجمالي' in item_name:
                        continue
                    cat_code  = detect_category(cat_text)
                    item_code = f"{cat_code}-{int(item_num):02d}"
                    rows.append({
                        'mosque_code':    mosque_code,
                        'item_code':      item_code,
                        'category':       cat_code,
                        'description':    item_name,
                        'uom':            map_uom(uom_text),
                        'contracted_qty': float(qty),
                        'unit_price':     float(price),
                    })
                except (ValueError, TypeError):
                    continue

        return self._process_rows(rows)

    # ── Core processor ────────────────────────────────────────────
    def _process_rows(self, rows):
        BOQ = self.env['mosque.boq']
        filter_codes = set(self.mosque_ids.mapped('code')) if self.mosque_ids else None

        created   = 0
        skipped   = 0
        not_found = set()
        by_mosque = {}

        for row in rows:
            mosque_code = row.get('mosque_code', '')
            if filter_codes and mosque_code not in filter_codes:
                continue

            mosque = self._get_mosque(mosque_code)
            if not mosque:
                not_found.add(mosque_code)
                skipped += 1
                continue

            if mosque.id not in by_mosque:
                by_mosque[mosque.id] = {'mosque': mosque, 'lines': []}
            by_mosque[mosque.id]['lines'].append(row)

        for mosque_id, data in by_mosque.items():
            mosque = data['mosque']
            if self.overwrite:
                BOQ.search([('mosque_id', '=', mosque.id)]).unlink()

            existing = set(BOQ.search(
                [('mosque_id', '=', mosque.id)]).mapped('item_code'))

            for row in data['lines']:
                item_code = row.get('item_code', '')
                if item_code in existing and not self.overwrite:
                    skipped += 1
                    continue

                cat = self._get_category(row.get('category', 'ARCH'))
                BOQ.create({
                    'mosque_id':      mosque.id,
                    'category_id':    cat.id,
                    'item_code':      item_code,
                    'description':    row.get('description', ''),
                    'uom':            row.get('uom', 'unit'),
                    'contracted_qty': row.get('contracted_qty', 0.0),
                    'unit_price':     row.get('unit_price', 0.0),
                })
                created += 1

        summary_lines = [
            f"✅ Created: {created} BOQ lines",
            f"⏭ Skipped: {skipped} lines",
        ]
        if not_found:
            summary_lines.append(
                f"⚠ Mosque codes not found: {', '.join(sorted(not_found))}")

        self.write({
            'result_summary': '\n'.join(summary_lines),
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_close(self):
        return {'type': 'ir.actions.act_window_close'}
