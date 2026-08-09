# -*- coding: utf-8 -*-
# controllers/ops_hub.py
from odoo import http
from odoo.http import request
from datetime import datetime, timedelta


class OpsHubController(http.Controller):

    def _get_portal_user(self):
        return request.env['waqf.portal.user'].sudo().search([
            ('user_id', '=', request.env.user.id),
            ('is_active', '=', True),
        ], limit=1) or None

    # ══════════════════════════════════════════════════════
    # MAIN HUB — /ops/hub
    # ══════════════════════════════════════════════════════
    @http.route('/ops/hub', type='http', auth='user', website=True)
    def ops_hub(self, tab=None, **kw):
        portal_user = self._get_portal_user()
        if portal_user:
            mosque_ids = portal_user.effective_mosque_ids.ids
        else:
            # مشرف الوقف — يرى كل المساجد
            mosque_ids = request.env['mosque.mosque'].sudo().search([]).ids

        if not mosque_ids:
            return request.redirect('/web')



        now = datetime.now()

        # ── وثائق ──────────────────────────────────────────
        Doc = request.env['waqf.document.approval'].sudo()
        doc_domain = ['|',
            ('submitted_by', '=', request.env.user.id),
            ('mosque_id', 'in', mosque_ids)]
        all_docs = Doc.search(doc_domain)

        doc_stats = {
            'total':    len(all_docs),
            'approved': len(all_docs.filtered(lambda d: d.state in ('approved', 'approved_comments'))),
            'pending':  len(all_docs.filtered(lambda d: d.state == 'submitted')),
            'rejected': len(all_docs.filtered(lambda d: d.state == 'rejected')),
            'draft':    len(all_docs.filtered(lambda d: d.state == 'draft')),
        }

        # ── عينات ──────────────────────────────────────────
        Sub = request.env['contractor.material.submittal'].sudo()
        all_subs = Sub.search([('mosque_id', 'in', mosque_ids)])

        sub_stats = {
            'total':    len(all_subs),
            'a': len(all_subs.filtered(lambda s: s.grade == 'a')),
            'b': len(all_subs.filtered(lambda s: s.grade == 'b')),
            'c': len(all_subs.filtered(lambda s: s.grade == 'c')),
            'd': len(all_subs.filtered(lambda s: s.grade == 'd')),
            'pending': len(all_subs.filtered(lambda s: s.state == 'submitted')),
        }

        # ── أوامر عمل ─────────────────────────────────────
        WO = request.env['contractor.work.order'].sudo()
        all_wos = WO.search([('mosque_id', 'in', mosque_ids)])

        wo_stats = {
            'total':     len(all_wos),
            'approved':  len(all_wos.filtered(lambda w: w.state == 'approved')),
            'delivered': len(all_wos.filtered(lambda w: w.state == 'delivered')),
            'closed':    len(all_wos.filtered(lambda w: w.state == 'closed')),
            'rework':    len(all_wos.filtered(lambda w: w.state == 'rework')),
        }

        # ── تأهيلات ────────────────────────────────────────
        Qual = request.env['contractor.qualification'].sudo()
        all_quals = Qual.search([
            '|',
            ('supervisor_id', '=', request.env.user.partner_id.id),
            ('mosque_ids', 'in', mosque_ids),
        ])

        qual_stats = {
            'total':    len(all_quals),
            'approved': len(all_quals.filtered(lambda q: q.state == 'approved')),
            'pending':  len(all_quals.filtered(
                lambda q: q.state in ('submitted', 'engineer_done', 'senior_done'))),
        }

        # ── إجمالي المعاملات ───────────────────────────────
        total_all = doc_stats['total'] + sub_stats['total'] + wo_stats['total'] + qual_stats['total']
        total_approved = doc_stats['approved'] + sub_stats['a'] + sub_stats['b'] + wo_stats['closed'] + qual_stats['approved']

        # ── حجم الملفات ────────────────────────────────────
        models_list = [
            'waqf.document.approval', 'waqf.document.approval.file',
            'contractor.material.submittal', 'contractor.work.order',
            'contractor.qualification',
        ]
        total_size = 0
        try:
            atts = request.env['ir.attachment'].sudo().search([
                ('res_model', 'in', models_list),
            ])
            total_size = sum(a.file_size or 0 for a in atts)
        except Exception:
            pass
        size_gb = total_size / (1024 ** 3)

        # ── متأخرات الاستشاري (معلّق > 3 أيام) ────────────
        three_days_ago = now - timedelta(days=3)
        delayed_docs = Doc.search([
            ('state', '=', 'submitted'),
            ('mosque_id', 'in', mosque_ids),
            ('write_date', '<', three_days_ago),
        ], order='write_date asc', limit=10)

        delayed_subs = Sub.search([
            ('state', '=', 'submitted'),
            ('mosque_id', 'in', mosque_ids),
            ('write_date', '<', three_days_ago),
        ], order='write_date asc', limit=10)

        # ── بانتظار الوقف ──────────────────────────────────
        pending_waqf_docs = Doc.search([
            ('state', '=', 'submitted'),
            ('mosque_id', 'in', mosque_ids),
        ], limit=10)
        pending_waqf_wos = WO.search([
            ('state', '=', 'delivered'),
            ('mosque_id', 'in', mosque_ids),
        ], limit=10)

        # عدد المتأخرات > 7 أيام
        seven_days_ago = now - timedelta(days=7)
        delayed_7_count = Doc.search_count([
            ('state', '=', 'submitted'),
            ('mosque_id', 'in', mosque_ids),
            ('write_date', '<', seven_days_ago),
        ]) + Sub.search_count([
            ('state', '=', 'submitted'),
            ('mosque_id', 'in', mosque_ids),
            ('write_date', '<', seven_days_ago),
        ])

        # ── أبرز أعمال المقاول (آخر 7 أيام) ───────────────
        week_ago = now - timedelta(days=7)
        recent_subs = Sub.search([
            ('mosque_id', 'in', mosque_ids),
            ('write_date', '>=', week_ago),
        ], order='write_date desc', limit=5)
        recent_docs = Doc.search([
            ('mosque_id', 'in', mosque_ids),
            ('write_date', '>=', week_ago),
        ], order='write_date desc', limit=5)
        recent_wos = WO.search([
            ('mosque_id', 'in', mosque_ids),
            ('write_date', '>=', week_ago),
        ], order='write_date desc', limit=5)

        # ── شجرة الملفات ───────────────────────────────────
        if portal_user:
            mosques = portal_user.effective_mosque_ids
        else:
            mosques = request.env['mosque.mosque'].sudo().search([])
        file_tree = []
        doc_types = request.env['waqf.document.type'].sudo().search([])

        for mosque in mosques:
            m_docs = Doc.search([('mosque_id', '=', mosque.id)])
            m_atts = request.env['ir.attachment'].sudo().search([
                ('res_model', '=', 'waqf.document.approval'),
                ('res_id', 'in', m_docs.ids),
            ])
            m_size = sum(a.file_size or 0 for a in m_atts)

            folders = []
            for dt in doc_types:
                dt_docs = m_docs.filtered(lambda d: d.doc_type_id.id == dt.id)
                if dt_docs:
                    dt_atts = request.env['ir.attachment'].sudo().search([
                        ('res_model', '=', 'waqf.document.approval'),
                        ('res_id', 'in', dt_docs.ids),
                    ])
                    folders.append({
                        'type': dt,
                        'docs': dt_docs,
                        'count': len(dt_docs),
                        'size': sum(a.file_size or 0 for a in dt_atts),
                    })

            file_tree.append({
                'mosque': mosque,
                'total_docs': len(m_docs),
                'total_size': m_size,
                'folders': folders,
            })

        # ── سجل الأحداث ───────────────────────────────────
        # نجمع من chatter كل النماذج
        activity_models = [
            'waqf.document.approval',
            'contractor.material.submittal',
            'contractor.work.order',
            'contractor.qualification',
        ]
        messages = request.env['mail.message'].sudo().search([
            ('model', 'in', activity_models),
            ('message_type', 'in', ('comment', 'notification')),
            ('subtype_id', '!=', False),
        ], order='date desc', limit=30)

        # ── الزيارات ───────────────────────────────────────
        visits = request.env['mosque.attendance'].sudo().search([
            ('mosque_id', 'in', mosque_ids),
            ('check_in', '>=', week_ago),
        ], order='check_in desc', limit=20)

        # ── إحصاءات الفترة ─────────────────────────────────
        month_ago = now - timedelta(days=30)
        month_visits = request.env['mosque.attendance'].sudo().search([
            ('mosque_id', 'in', mosque_ids),
            ('check_in', '>=', month_ago),
        ])
        avg_duration = 0
        if month_visits:
            durations = [v.duration for v in month_visits if v.duration]
            avg_duration = sum(durations) / len(durations) if durations else 0

        # متوسط وقت المعالجة (وثائق)
        approved_docs = Doc.search([
            ('mosque_id', 'in', mosque_ids),
            ('state', 'in', ('approved', 'approved_comments')),
            ('write_date', '>=', month_ago),
        ])
        avg_process = 0
        if approved_docs:
            deltas = []
            for d in approved_docs:
                if d.create_date and d.write_date:
                    delta = (d.write_date - d.create_date).total_seconds() / 86400
                    deltas.append(delta)
            avg_process = sum(deltas) / len(deltas) if deltas else 0

        approval_rate = 0
        if total_all:
            approval_rate = (total_approved / total_all) * 100

        return request.render('waqf_contractor_portal.tmpl_ops_hub', {
            'portal_user':    portal_user,
            'active_tab':     tab or 'overview',
            # KPIs
            'total_all':      total_all,
            'total_approved': total_approved,
            'total_pending':  doc_stats['pending'] + sub_stats['pending'],
            'delayed_7':      delayed_7_count,
            'size_gb':        round(size_gb, 1),
            # Stats
            'doc_stats':      doc_stats,
            'sub_stats':      sub_stats,
            'wo_stats':       wo_stats,
            'qual_stats':     qual_stats,
            # Alerts
            'delayed_docs':   delayed_docs,
            'delayed_subs':   delayed_subs,
            'pending_waqf_docs': pending_waqf_docs,
            'pending_waqf_wos':  pending_waqf_wos,
            # Recent
            'recent_subs':    recent_subs,
            'recent_docs':    recent_docs,
            'recent_wos':     recent_wos,
            # File tree
            'file_tree':      file_tree,
            # Activity log
            'messages':       messages,
            'visits':         visits,
            # Period stats
            'month_visits':   len(month_visits),
            'avg_duration':   round(avg_duration, 1),
            'avg_process':    round(avg_process, 1),
            'approval_rate':  round(approval_rate),
            # Helpers
            'now':            now,
            'format_size':    lambda s: '%.1f MB' % (s / 1048576) if s else '0',
            'days_since':     lambda d: (now - d).days if d else 0,
        })

    # ══════════════════════════════════════════════════════
    # EXPORT EXCEL
    # ══════════════════════════════════════════════════════
    @http.route('/ops/hub/export/excel', type='http', auth='user', website=True)
    def ops_export_excel(self, **kw):
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return request.redirect('/ops/hub?error=openpyxl')

        portal_user = self._get_portal_user()
        if not portal_user:
            return request.redirect('/web')

        mosque_ids = portal_user.effective_mosque_ids.ids
        wb = openpyxl.Workbook()

        hdr_font = Font(name='Cairo', bold=True, size=11, color='FFFFFF')
        hdr_fill = PatternFill(start_color='0F1B2D', end_color='0F1B2D', fill_type='solid')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_font = Font(name='Cairo', size=10)
        thin = Border(
            left=Side(style='thin', color='E2E8ED'),
            right=Side(style='thin', color='E2E8ED'),
            top=Side(style='thin', color='E2E8ED'),
            bottom=Side(style='thin', color='E2E8ED'))

        def write_headers(ws, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = hdr_align; c.border = thin

        # ── Sheet 1: وثائق ─────────────────────────────────
        ws1 = wb.active
        ws1.title = 'الوثائق'
        ws1.sheet_view.rightToLeft = True
        write_headers(ws1, ['الرقم', 'العنوان', 'النوع', 'المسجد', 'الحالة', 'تاريخ التقديم', 'المقدّم'])
        docs = request.env['waqf.document.approval'].sudo().search([
            ('mosque_id', 'in', mosque_ids)], order='create_date desc')
        state_map = {'draft': 'مسودة', 'submitted': 'بانتظار', 'approved': 'معتمد',
                     'approved_comments': 'معتمد بملاحظات', 'rejected': 'مرفوض'}
        for i, d in enumerate(docs, 2):
            vals = [d.name or '', d.title or '', d.doc_type_id.name if d.doc_type_id else '',
                    d.mosque_id.name if d.mosque_id else '', state_map.get(d.state, d.state or ''),
                    str(d.create_date or '')[:16], d.submitted_by.name if d.submitted_by else '']
            for col, v in enumerate(vals, 1):
                c = ws1.cell(row=i, column=col, value=v)
                c.font = data_font; c.border = thin

        # ── Sheet 2: عينات ─────────────────────────────────
        ws2 = wb.create_sheet('العينات')
        ws2.sheet_view.rightToLeft = True
        write_headers(ws2, ['الرقم', 'المادة', 'المصنع', 'المسجد', 'التقييم', 'الإصدار',
                            'التاريخ', 'المراجع', 'ملاحظات الاستشاري'])
        subs = request.env['contractor.material.submittal'].sudo().search([
            ('mosque_id', 'in', mosque_ids)], order='date_submitted desc')
        for i, s in enumerate(subs, 2):
            vals = [s.name or '', s.material_name or '', s.manufacturer or '',
                    s.mosque_id.name if s.mosque_id else '', (s.grade or '').upper(),
                    s.revision or 0, str(s.date_submitted or ''),
                    s.reviewed_by.name if s.reviewed_by else '', s.review_notes or '']
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row=i, column=col, value=v)
                c.font = data_font; c.border = thin

        # ── Sheet 3: أوامر عمل ─────────────────────────────
        ws3 = wb.create_sheet('أوامر العمل')
        ws3.sheet_view.rightToLeft = True
        write_headers(ws3, ['الرقم', 'الوصف', 'المسجد', 'الحالة', 'التقييم', 'القيمة',
                            'تاريخ الطلب'])
        wos = request.env['contractor.work.order'].sudo().search([
            ('mosque_id', 'in', mosque_ids)], order='date_requested desc')
        for i, w in enumerate(wos, 2):
            vals = [w.name or '', w.work_description or '',
                    w.mosque_id.name if w.mosque_id else '', w.state or '',
                    (w.grade or '').upper(), w.total_value or 0,
                    str(w.date_requested or '')]
            for col, v in enumerate(vals, 1):
                c = ws3.cell(row=i, column=col, value=v)
                c.font = data_font; c.border = thin

        # ── Sheet 4: زيارات ────────────────────────────────
        ws4 = wb.create_sheet('الزيارات')
        ws4.sheet_view.rightToLeft = True
        write_headers(ws4, ['المسجد', 'المهندس', 'الدخول', 'الخروج', 'المدة (ساعة)', 'المسافة (م)'])
        month_ago = datetime.now() - timedelta(days=30)
        visits = request.env['mosque.attendance'].sudo().search([
            ('mosque_id', 'in', mosque_ids),
            ('check_in', '>=', month_ago),
        ], order='check_in desc')
        for i, v in enumerate(visits, 2):
            eng_name = ''
            if v.engineer_id:
                eng_name = v.engineer_id.name
            elif v.portal_user_id:
                eng_name = v.portal_user_id.name
            vals = [v.mosque_id.name if v.mosque_id else '', eng_name,
                    str(v.check_in or '')[:16], str(v.check_out or '')[:16],
                    round(v.duration or 0, 1), round(v.distance_m or 0)]
            for col, val in enumerate(vals, 1):
                c = ws4.cell(row=i, column=col, value=val)
                c.font = data_font; c.border = thin

        # ── حفظ ────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        fname = 'ops_report_%s.xlsx' % datetime.now().strftime('%Y%m%d')

        return request.make_response(output.read(), headers=[
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', 'attachment; filename="%s"' % fname),
        ])
